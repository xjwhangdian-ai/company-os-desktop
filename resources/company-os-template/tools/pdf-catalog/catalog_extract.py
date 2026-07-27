#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
产品画册 PDF → 抽图 + 文本提取 + 配对定稿（供 sales 分身/人工整理成产品清单用）。

两个阶段：
  阶段1 抽取：python3 catalog_extract.py <画册.pdf> <产出目录>
    _中间产物/候选图/  每页检测到的候选产品照（P{页}_{序}.jpg）——数量偏多是正常的
    _中间产物/整页/    每页整页位图，配对/复核时看整页
    _中间产物/标注/    每页标注候选框编号的图（原图1/2大小），分身照它核对"编号↔产品"
    _中间产物/候选.json  结构化候选框坐标
    文本提取.md        每页文本层（型号/名称/参数原文）

  阶段2 定稿：python3 catalog_extract.py --apply <产出目录> [--clean]
    --clean：定稿成功后删除 _中间产物/（可随时重新抽取再生），最终目录只留
             产品图片/ + 文本提取.md + _配对.json
    读分身写的 <产出目录>/_配对.json：
      [{"序号":1,"型号":"HW-1","产品名称":"排爆服","页":5,"候选":0}, {"...","框":[x0,y0,x1,y1]}]
      候选=阶段1候选编号；候选不合适时用 框（按 _整页/P{页}.jpg 原图像素）。
    产出 <产出目录>/产品图片/：每个产品一张成品图，命名 序号_型号_产品名称_P页.jpg。

说明：数字排版画册效果最好；整页扫描件抽不出独立图。一页多产品的精确图文配对由
sales 分身照 _标注/ 逐页核对后写 _配对.json（配图铁律），机械只负责抠图与执行清单。
依赖：pypdf、Pillow、numpy（缺失时 App 会提示 pip 安装）。
"""
import sys, os, json, re, warnings
warnings.filterwarnings("ignore")

try:
    from pypdf import PdfReader
    from PIL import Image, ImageFilter, ImageDraw, ImageFont
    import numpy as np
except ImportError as e:
    print("MISSING_DEP:" + str(e), file=sys.stderr)
    sys.exit(3)


def page_bitmap(page):
    """取该页最大的嵌入位图（数字画册每页通常一张整页底图）。"""
    imgs = list(page.images)
    if not imgs:
        return None
    main = max(imgs, key=lambda im: len(im.data))
    import io
    try:
        return Image.open(io.BytesIO(main.data)).convert("RGB")
    except Exception:
        return None


def detect_boxes(im, min_area_frac=0.004):
    """在整页位图上找"产品照候选块"：饱和/暗的密集连通域，过滤扁长条(标题栏)与稀碎文字。"""
    W, H = im.size
    k = 4
    sm = im.resize((max(1, W // k), max(1, H // k)))
    a = np.asarray(sm).astype(int)
    r, g, b = a[..., 0], a[..., 1], a[..., 2]
    mx = np.maximum(np.maximum(r, g), b)
    mn = np.minimum(np.minimum(r, g), b)
    sat = mx - mn
    fg = (mx < 180) | (sat > 40)
    m = np.asarray(Image.fromarray((fg * 255).astype("uint8")).filter(ImageFilter.MaxFilter(7))) > 0
    h, w = m.shape
    seen = np.zeros_like(m, bool)
    out = []
    for y in range(h):
        for x in range(w):
            if m[y, x] and not seen[y, x]:
                st = [(y, x)]
                seen[y, x] = True
                n = 0
                x0 = x1 = x
                y0 = y1 = y
                while st:
                    cy, cx = st.pop()
                    n += 1
                    x0 = min(x0, cx); x1 = max(x1, cx); y0 = min(y0, cy); y1 = max(y1, cy)
                    for dy, dx in ((1, 0), (-1, 0), (0, 1), (0, -1)):
                        ny, nx = cy + dy, cx + dx
                        if 0 <= ny < h and 0 <= nx < w and m[ny, nx] and not seen[ny, nx]:
                            seen[ny, nx] = True
                            st.append((ny, nx))
                bw, bh = (x1 - x0 + 1), (y1 - y0 + 1)
                area = bw * bh
                if area < h * w * min_area_frac:
                    continue
                ar = bw / bh
                fill = n / area
                if ar > 4 or ar < 0.15 or fill < 0.25:
                    continue
                if bh * k < 160 or bw * k < 120:  # 过小的块多为 logo/二维码/小图标
                    continue
                out.append(dict(x0=x0 * k, y0=y0 * k, x1=x1 * k, y1=y1 * k,
                                w=bw * k, h=bh * k, fill=round(fill, 2)))
    # 左栏在前、上到下，编号稳定
    out.sort(key=lambda bx: (0 if (bx["x0"] + bx["x1"]) / 2 < W / 2 else 1, bx["y0"]))
    return out


def _slug(s):
    return re.sub(r'[\\/:*?"<>|\s]+', "_", str(s or "")).strip("_")[:36]



def page_text_runs(page):
    """带坐标/字号的文本片段（PDF 用户空间坐标，y 原点在页底）；部分 PDF 的 XObject 文本取不到。"""
    runs = []

    def visitor(text, cm, tm, font_dict, font_size):
        s = (text or "").strip()
        if not s:
            return
        try:
            eff = abs((font_size or 0) * (tm[3] if tm[3] else 1))
        except Exception:
            eff = font_size or 0
        runs.append((s, float(tm[4]), float(tm[5]), float(eff)))

    try:
        page.extract_text(visitor_text=visitor)
    except Exception:
        pass
    return runs


def _ocr_lines(img_path):
    """macOS 系统 Vision OCR（离线）；不可用返回 None。返回 [(text, x_img, y_top_img, h_img), ...]"""
    try:
        from ocrmac import ocrmac as _o
    except Exception:
        return None
    try:
        with Image.open(img_path) as im:
            iw, ih = im.size
        out = []
        for txt, _conf, box in _o.OCR(img_path, language_preference=["zh-Hans", "en-US"]).recognize():
            s = (txt or "").strip()
            if not s:
                continue
            x, y, w, h = box  # 归一化，原点左下
            out.append((s, x * iw, (1 - (y + h)) * ih, h * ih))
        return out
    except Exception:
        return None


_MODELISH = re.compile(r"^[A-Za-z0-9][A-Za-z0-9\-/\.+()®™ ]{1,}$")
# 前言/栏目类标题不是产品名（公司简介、资质荣誉、技术参数小节标题等）
_JUNK_NAME = re.compile(r"简介|文化|荣誉|目录|联系|扫码|愿景|理念|资质|租赁|回收|服务商|供货单位|一站式|技术参数|产品特点|产品参数|配置清单|规格参数|技术规格|解决方案|应用场景|公司|集团")
# 产品页信号：正文含这些词的页才认为在介绍具体产品
_PRODUCT_PAGE = re.compile(r"产品简介|技术参数|产品特点|规格参数|技术规格|产品参数|配置清单")


def _lines_to_headings(lines, iw, ih):
    """大字行 → 标题块（图像坐标）。lines: [(text, x_img, y_top_img, h_img)]"""
    body = sorted([l[3] for l in lines if len(l[0]) >= 2 and l[3] > 0])
    if not body:
        return []
    med = body[len(body) // 2]
    big = []
    for s, x, y_top, h in lines:
        if h < med * 1.4 or len(s) < 2:
            continue
        if y_top < ih * 0.055 or y_top > ih * 0.95:  # 页眉/页脚
            continue
        cjk = len(re.findall(r"[\u4e00-\u9fa5]", s))
        modelish = cjk == 0 and _MODELISH.match(s) and any(c.isdigit() or c == "-" for c in s)
        if cjk < 2 and not modelish:
            continue
        big.append((s, x, y_top, h))
    big.sort(key=lambda l: (0 if l[1] < iw / 2 else 1, l[2]))
    headings = []
    for s, x, y_top, h in big:
        col = 0 if x < iw / 2 else 1
        if headings and headings[-1]["col"] == col and abs(y_top - headings[-1]["y_last"]) < h * 2.8:
            headings[-1]["lines"].append(s)
            headings[-1]["y_last"] = y_top
        else:
            headings.append({"col": col, "y_top": y_top, "y_last": y_top, "lines": [s]})
    return headings


def _collect_page(pg, boxes, headings, iw, ih, seq_start, body_lines=None):
    """在一页内按标题块认领候选图，返回该页 items（序号从 seq_start+1 起）。
    body_lines: 该页全部文本行 [(text, x, y_top, h)]，用于收集产品地盘内的正文作技术参数。"""
    items = []
    seq = seq_start
    used = set()
    for hi, h in enumerate(headings):
        model, name = "", ""
        for ln in h["lines"]:
            cjk = len(re.findall(r"[\u4e00-\u9fa5]", ln))
            if not model and cjk == 0 and _MODELISH.match(ln) and any(c.isdigit() or c == "-" for c in ln):
                model = ln.strip()
            elif not name and cjk >= 2:
                name = ln.strip()
        if not (model or name):
            continue
        if not model and name and _JUNK_NAME.search(name):
            continue  # 无型号且名称是栏目词 → 不是产品
        nxt = None
        for h2 in headings[hi + 1:]:
            if h2["col"] == h["col"]:
                nxt = h2["y_top"]
                break
        y_bot = nxt if nxt is not None else ih
        best, best_area = None, 0
        for bi, b in enumerate(boxes):
            if bi in used:
                continue
            bcx = (b["x0"] + b["x1"]) / 2
            bcy = (b["y0"] + b["y1"]) / 2
            if (0 if bcx < iw / 2 else 1) != h["col"]:
                continue
            if not (h["y_top"] - 60 <= bcy <= y_bot + 40):
                continue
            area = b["w"] * b["h"]
            if area > best_area:
                best, best_area = bi, area
        if best is None:
            continue
        used.add(best)
        seq += 1
        # 技术参数：同栏、标题以下到下一标题之间的正文行（排除标题自身），按纵向顺序拼接
        params = ""
        if body_lines:
            own = set(h["lines"])
            zone = []
            for s, x, y_top, lh in body_lines:
                if s in own or len(s.strip()) < 2:
                    continue
                if (0 if x < iw / 2 else 1) != h["col"]:
                    continue
                if not (h["y_last"] + lh * 0.5 <= y_top <= y_bot - 2):
                    continue
                zone.append((y_top, x, s.strip()))
            zone.sort()
            params = "\n".join(s for _, _, s in zone)[:1200]
        items.append({"序号": seq, "型号": model, "产品名称": name, "分类": "", "页": pg, "候选": best,
                      "技术参数": params})
    return items


def auto_pair(reader, meta, out):
    """机械自动配对（不依赖 AI）：每页找型号/名称标题块（文本层优先，扫描页用系统 OCR），
    候选图按同栏纵向地盘就近认领；只对"产品页"（正文含 产品简介/技术参数 等）配对。
    全部失败时兜底导出全部候选（名称留空）。"""
    pages = []  # (pg, boxes, headings, iw, ih, is_product_page)
    any_heading = False
    used_ocr = False
    for m in meta:
        pg = m["page"]
        boxes = m.get("boxes") or []
        if not boxes:
            continue
        pp = os.path.join(out, "_中间产物", "整页", "P%02d.jpg" % pg)
        try:
            with Image.open(pp) as pim:
                iw, ih = pim.size
        except Exception:
            continue
        # 1) 文本层（转图像坐标）
        page = reader.pages[pg - 1]
        runs = page_text_runs(page)
        lines = []
        if runs:
            mb = page.mediabox
            try:
                pw, ph = float(mb.width), float(mb.height)
            except Exception:
                pw, ph = 595.0, 842.0
            sx, sy = iw / pw, ih / ph
            lines = [(s, x * sx, (ph - y) * sy - fs * sy, fs * sy) for (s, x, y, fs) in runs]
        headings = _lines_to_headings(lines, iw, ih) if lines else []
        # 2) 扫描页兜底：系统 OCR
        ocr_lines_cache = None
        if not headings:
            ocr_lines_cache = _ocr_lines(pp)
            if ocr_lines_cache:
                used_ocr = True
                headings = _lines_to_headings(ocr_lines_cache, iw, ih)
        if not headings:
            continue
        any_heading = True
        all_lines = (lines or []) + (ocr_lines_cache or [])
        all_text = " ".join(l[0] for l in all_lines)
        pages.append((pg, boxes, headings, iw, ih, bool(_PRODUCT_PAGE.search(all_text)), all_lines))
    # 第一轮：只配产品页；一个产品页都没有（版式特殊）再放开
    items = []
    seq = 0
    for gate in (True, False):
        items = []
        seq = 0
        for pg, boxes, headings, iw, ih, is_prod, all_lines in pages:
            if gate and not is_prod:
                continue
            got = _collect_page(pg, boxes, headings, iw, ih, seq, all_lines)
            items.extend(got)
            seq += len(got)
        if items:
            break
    if not any_heading:
        items = []
        seq = 0
        for m in meta:
            for bi, _b in enumerate(m.get("boxes") or []):
                seq += 1
                items.append({"序号": seq, "型号": "", "产品名称": "", "分类": "", "页": m["page"], "候选": bi})
        return items, True, used_ocr
    return items, False, used_ocr


def _save_final(im, dst, min_side=800):
    """成品图保存：两边都要 ≥ min_side，不足则等比放大（LANCZOS）；只放大不缩小。"""
    w, h = im.size
    scale = max(min_side / w, min_side / h)
    if scale > 1:
        im = im.resize((max(min_side, round(w * scale)), max(min_side, round(h * scale))), Image.LANCZOS)
    im.save(dst, quality=92)


_SHEET_COLS = ["序号", "产品名称", "产品类别", "品牌", "型号 / 规格", "生产制造商", "产地",
               "技术规格参数", "单位", "税率", "质保期(月)", "供货价(含税单价/元)", "交货周期(天)",
               "最小起订量", "开票类型", "备注", "对应画册抠图的序列"]


def derive_brand(pdf_name):
    """从画册文件名推断品牌/生产制造商（如 和为永泰产品画册-2026.pdf → 和为永泰）。"""
    base = os.path.splitext(os.path.basename(pdf_name))[0]
    base = re.sub(r"产品画册|产品手册|画册|手册|目录|宣传册|\d{4}年?版?|20\d{2}|_压缩|[-_（(【\[].*$", "", base)
    base = re.sub(r"[-_\s]+$", "", base).strip()
    return base if 2 <= len(base) <= 20 else ""


def _write_product_sheet(out, rows):
    """按《供应商资料解析模板》格式产出 产品清单.xlsx：
    末列直接嵌入对应抠图（原图≥800×800，显示缩放）；型号缺失的单元格标黄（必填待补）。
    openpyxl 不在时降级 CSV（图片列写文件名）。"""
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "供应商报价清单"
        ws.cell(row=1, column=1, value="供应商报价清单").font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_SHEET_COLS))
        ws.cell(row=2, column=1, value="名称/型号/技术参数/品牌/制造商由画册抠图机械提取（个别字可能有误差，请对照抠图核对）；型号为必填项，黄色=画册中未识别到、需人工补；价格/税率等空列请人工或供应商补充。抠图原图分辨率≥800×800，表内为缩放显示。")
        head_fill = PatternFill("solid", fgColor="E8EEF6")
        model_missing_fill = PatternFill("solid", fgColor="FFF2A8")
        for ci, name in enumerate(_SHEET_COLS, 1):
            c = ws.cell(row=3, column=ci, value=name)
            c.font = Font(bold=True)
            c.fill = head_fill
        widths = [6, 22, 14, 12, 18, 14, 8, 50, 6, 8, 10, 16, 12, 10, 10, 12, 20]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=3, column=ci).column_letter].width = w
        img_col_letter = ws.cell(row=3, column=len(_SHEET_COLS)).column_letter
        DISPLAY = 130  # 图片显示尺寸（px）；嵌入的原图文件仍是 ≥800×800
        for ri, r in enumerate(rows, 4):
            for ci, name in enumerate(_SHEET_COLS, 1):
                if name == "对应画册抠图的序列":
                    continue
                cell = ws.cell(row=ri, column=ci, value=r.get(name, ""))
                if name == "技术规格参数":
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                elif name == "型号 / 规格" and not r.get(name):
                    cell.fill = model_missing_fill
            img_path = r.get("图片路径") or ""
            if img_path and os.path.exists(img_path):
                try:
                    img = XLImage(img_path)
                    iw, ih = img.width or DISPLAY, img.height or DISPLAY
                    scale = DISPLAY / max(iw, ih)
                    img.width, img.height = round(iw * scale), round(ih * scale)
                    ws.add_image(img, "%s%d" % (img_col_letter, ri))
                    ws.row_dimensions[ri].height = DISPLAY * 0.75 + 6  # px→pt
                except Exception:
                    ws.cell(row=ri, column=len(_SHEET_COLS), value=os.path.basename(img_path))
            else:
                ws.cell(row=ri, column=len(_SHEET_COLS), value=r.get("对应画册抠图的序列", ""))
        p = os.path.join(out, "产品清单.xlsx")
        wb.save(p)
        return p
    except ImportError:
        import csv
        p = os.path.join(out, "产品清单.csv")
        with open(p, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=_SHEET_COLS)
            w.writeheader()
            for r in rows:
                w.writerow({k: r.get(k, "") for k in _SHEET_COLS})
        return p


def apply_pairing(out, clean=False):
    """阶段2：按分身写的 _配对.json 产出成品图 产品图片/序号_型号_产品名称_P页.jpg"""
    import shutil
    pairing = os.path.join(out, "_配对.json")
    if not os.path.exists(pairing):
        print(json.dumps({"ok": False, "needPairing": True,
                          "说明": "还没有 _配对.json——先让分身照 _标注/ 核对并生成配对清单"}, ensure_ascii=False))
        return
    try:
        items = json.load(open(pairing, encoding="utf-8"))
        assert isinstance(items, list) and items
    except Exception as e:
        print(json.dumps({"ok": False, "needPairing": True,
                          "说明": "_配对.json 解析失败（%s）——请让分身重新生成" % e}, ensure_ascii=False))
        return
    dest = os.path.join(out, "产品图片")
    os.makedirs(dest, exist_ok=True)
    # 目录布局兼容：新 _中间产物/候选图 → 旧 _候选图 → 更旧 产品图候选
    cand_dir = None
    for c in ("_中间产物/候选图", "_候选图", "产品图候选"):
        p = os.path.join(out, c)
        if os.path.isdir(p):
            cand_dir = p
            break
    page_dir = None
    for c in ("_中间产物/整页", "_整页"):
        p = os.path.join(out, c)
        if os.path.isdir(p):
            page_dir = p
            break
    # 同画册品牌/生产制造商自动补全（来源：抽取阶段写的 _画册信息.json，从画册文件名推断）
    brand = ""
    try:
        info = json.load(open(os.path.join(out, "_画册信息.json"), encoding="utf-8"))
        brand = str(info.get("品牌") or "")
    except Exception:
        pass
    okn, miss = 0, []
    sheet_rows = []
    for it in items:
        try:
            seq = int(it.get("序号"))
            pg = int(it.get("页"))
        except Exception:
            miss.append(str(it.get("产品名称") or it.get("型号") or "?"))
            continue
        parts = ["%03d" % seq, _slug(it.get("型号")), _slug(it.get("产品名称")), "P%02d" % pg]
        fn = "_".join([p for p in parts if p]) + ".jpg"
        sheet_rows.append({
            "序号": seq,
            "产品名称": it.get("产品名称") or "",
            "产品类别": it.get("分类") or "",
            "品牌": brand,
            "生产制造商": brand,
            "型号 / 规格": it.get("型号") or "",
            "技术规格参数": it.get("技术参数") or "",
            "备注": "手册P%02d" % pg,
            "对应画册抠图的序列": fn,
            "图片路径": os.path.join(dest, fn)
        })
        dst = os.path.join(dest, fn)
        cand = it.get("候选")
        box = it.get("框")
        done = False
        if cand is not None and cand_dir:
            src = os.path.join(cand_dir, "P%02d_%d.jpg" % (pg, int(cand)))
            if os.path.exists(src):
                _save_final(Image.open(src).convert("RGB"), dst)
                done = True
        if not done and box and len(box) == 4:
            page_img = os.path.join(page_dir or os.path.join(out, "_整页"), "P%02d.jpg" % pg)
            if os.path.exists(page_img):
                im = Image.open(page_img).convert("RGB")
                x0, y0, x1, y1 = [int(v) for v in box]
                x0, y0 = max(0, x0), max(0, y0)
                x1, y1 = min(im.size[0], x1), min(im.size[1], y1)
                if x1 - x0 > 20 and y1 - y0 > 20:
                    _save_final(im.crop((x0, y0, x1, y1)), dst)
                    done = True
        if done:
            okn += 1
        else:
            miss.append("%s(P%d 候选%s)" % (it.get("产品名称") or it.get("型号") or "?", pg, cand))
    # 产品清单 Excel（模板同款 17 列：名称/型号/技术参数机械填入，其余留空待补）
    sheet_path = ""
    if sheet_rows:
        try:
            sheet_path = _write_product_sheet(out, sheet_rows)
        except Exception:
            sheet_path = ""
    cleaned = False
    if clean and okn > 0:
        # 定稿成功后清理可再生的中间产物，最终目录只留 产品图片/ + 文本提取.md + _配对.json
        for c in ("_中间产物", "_候选图", "_整页", "_标注", "产品图候选", "_候选.json", "_核对进度.json"):
            p = os.path.join(out, c)
            try:
                if os.path.isdir(p):
                    shutil.rmtree(p)
                    cleaned = True
                elif os.path.isfile(p):
                    os.remove(p)
                    cleaned = True
            except Exception:
                pass
    print(json.dumps({"ok": True, "count": okn, "missing": miss, "outDir": dest, "cleaned": cleaned,
                      "sheet": sheet_path,
                      "说明": "已按配对清单产出 %d 张成品图到 产品图片/（命名 序号_型号_产品名称_P页）%s%s%s"
                              % (okn,
                                 ("＋产品清单.%s（名称/型号/技术参数已填，价格等待补）" % ("xlsx" if sheet_path.endswith(".xlsx") else "csv")) if sheet_path else "",
                                 ("；%d 条未命中：%s" % (len(miss), "、".join(miss[:5]))) if miss else "",
                                 "；中间产物已清理" if cleaned else "")},
                     ensure_ascii=False))


def main():
    if len(sys.argv) >= 3 and sys.argv[1] == "--apply":
        apply_pairing(sys.argv[2], clean="--clean" in sys.argv[3:])
        return
    if len(sys.argv) < 3:
        print("用法: catalog_extract.py <画册.pdf> <产出目录> | --apply <产出目录> [--clean]", file=sys.stderr)
        sys.exit(1)
    pdf, out = sys.argv[1], sys.argv[2]
    mid = os.path.join(out, "_中间产物")
    d_crop = os.path.join(mid, "候选图")
    d_page = os.path.join(mid, "整页")
    d_mark = os.path.join(mid, "标注")
    for d in (d_crop, d_page, d_mark):
        os.makedirs(d, exist_ok=True)

    reader = PdfReader(pdf)
    font = None
    for fp in ("/System/Library/Fonts/Helvetica.ttc",          # macOS
               "C:/Windows/Fonts/arial.ttf",                    # Windows
               "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"):  # Linux 兜底
        try:
            font = ImageFont.truetype(fp, 64)
            break
        except Exception:
            continue
    if font is None:
        font = ImageFont.load_default()

    md = ["# %s · 文本提取\n" % os.path.basename(pdf),
          "> 每页文本层原文（型号/名称/参数），配合 _标注/ 里的候选框编号核对配图。\n"]
    meta = []
    n_crop = 0
    for i, page in enumerate(reader.pages):
        pno = i + 1
        text = (page.extract_text() or "").strip()
        im = page_bitmap(page)
        boxes = []
        if im is not None:
            im.save(os.path.join(d_page, "P%02d.jpg" % pno), quality=85)
            boxes = detect_boxes(im)
            mark = im.copy()
            dr = ImageDraw.Draw(mark)
            for idx, b in enumerate(boxes):
                pad = 12
                crop = im.crop((max(0, b["x0"] - pad), max(0, b["y0"] - pad),
                                min(im.size[0], b["x1"] + pad), min(im.size[1], b["y1"] + pad)))
                crop.save(os.path.join(d_crop, "P%02d_%d.jpg" % (pno, idx)), quality=90)
                n_crop += 1
                dr.rectangle([b["x0"], b["y0"], b["x1"], b["y1"]], outline=(255, 0, 0), width=8)
                dr.rectangle([b["x0"], b["y0"] - 74, b["x0"] + 96, b["y0"]], fill=(255, 0, 0))
                dr.text((b["x0"] + 12, b["y0"] - 72), str(idx), fill=(255, 255, 255), font=font)
            mark.resize((mark.size[0] // 2, mark.size[1] // 2)).save(
                os.path.join(d_mark, "P%02d.jpg" % pno), quality=80)
        if text:
            md.append("## 第 %d 页\n\n%s\n" % (pno, text))
        meta.append(dict(page=pno, text=text, boxes=boxes))

    with open(os.path.join(out, "文本提取.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(md))
    with open(os.path.join(mid, "候选.json"), "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False)
    # 画册信息（品牌从文件名推断，供 --apply 阶段补全 品牌/生产制造商）
    with open(os.path.join(out, "_画册信息.json"), "w", encoding="utf-8") as f:
        json.dump({"来源PDF": os.path.basename(pdf), "品牌": derive_brand(pdf)}, f, ensure_ascii=False)

    # 机械自动配对（不依赖 AI）：文本层大字标题 ↔ 候选图就近认领；扫描版兜底导出全部候选
    pair_items, degraded, used_ocr = auto_pair(reader, meta, out)
    if pair_items:
        with open(os.path.join(out, "_配对.json"), "w", encoding="utf-8") as f:
            json.dump(pair_items, f, ensure_ascii=False, indent=1)

    print(json.dumps({
        "ok": True, "pages": len(reader.pages), "crops": n_crop,
        "autoPaired": len(pair_items), "degraded": degraded, "usedOcr": used_ocr,
        "outDir": out,
        "说明": "阶段1完成：抽出 %d 张候选图（_候选图/，中间产物）+ 逐页文本 + 标注图；"
                "下一步由分身照 _标注/ 核对生成 _配对.json，再执行定稿出成品图" % n_crop
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
